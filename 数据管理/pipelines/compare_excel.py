#!/usr/bin/env python3
import argparse
import json
import math
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter


_SHEET_SAFE_RE = re.compile(r"[^0-9A-Za-z\u4e00-\u9fff _\-\(\)\[\]\.]+")


def _now_tag() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _to_list(x: Optional[str]) -> List[str]:
    if not x:
        return []
    return [s.strip() for s in x.split(",") if s.strip()]


def _coerce_sheet_arg(sheet: str):
    s = (sheet or "").strip()
    if not s or s.lower() == "all":
        return "all"
    if s.isdigit():
        return int(s)
    return s


def _normalize_columns(cols: Sequence[str]) -> List[str]:
    out = []
    for c in cols:
        c2 = str(c).strip()
        out.append(c2)
    return out


def _safe_sheet_name(name: str, suffix: str = "") -> str:
    n = _SHEET_SAFE_RE.sub("_", str(name).strip())
    n = n[:31]
    if suffix:
        base = n[: max(0, 31 - len(suffix))]
        n = f"{base}{suffix}"
    if not n:
        n = "Sheet"
    return n[:31]


def _escape_html(x: object) -> str:
    s = "" if x is None else str(x)
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def _auto_columns_width(ws, df: pd.DataFrame, max_width: int = 60):
    for i, col in enumerate(df.columns, start=1):
        if i > 16384:
            break
        values = df[col].astype(str).fillna("")
        m = max([len(str(col))] + values.head(2000).map(len).tolist())
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, m + 2), max_width)


def _read_excel(path: str, sheet, usecols: Optional[List[str]]):
    if usecols:
        return pd.read_excel(path, sheet_name=sheet, dtype=object, usecols=usecols)
    return pd.read_excel(path, sheet_name=sheet, dtype=object)


def _as_string_series(s: pd.Series, trim: bool) -> pd.Series:
    out = s.copy()
    out = out.where(~out.isna(), None)
    out = out.map(lambda x: "" if x is None else str(x))
    if trim:
        out = out.map(lambda x: x.strip())
    return out


def _is_numeric_like(a: pd.Series, b: pd.Series, threshold: float) -> bool:
    an = pd.to_numeric(a, errors="coerce")
    bn = pd.to_numeric(b, errors="coerce")
    ra = float(an.notna().mean()) if len(an) else 0.0
    rb = float(bn.notna().mean()) if len(bn) else 0.0
    return min(ra, rb) >= threshold


def _numeric_diff_mask(a: pd.Series, b: pd.Series, tol: float) -> pd.Series:
    an = pd.to_numeric(a, errors="coerce")
    bn = pd.to_numeric(b, errors="coerce")
    both_nan = an.isna() & bn.isna()
    diff = (an - bn).abs()
    mask = ~both_nan & (diff > tol)
    return mask.fillna(False)


def _string_diff_mask(a: pd.Series, b: pd.Series, trim: bool) -> pd.Series:
    sa = _as_string_series(a, trim=trim)
    sb = _as_string_series(b, trim=trim)
    return sa.ne(sb)


def _ensure_unique_index(df: pd.DataFrame, keys: List[str], allow_dup: bool) -> Tuple[pd.DataFrame, List[str]]:
    if not keys:
        df2 = df.copy()
        df2["__row__"] = np.arange(len(df2), dtype=np.int64) + 1
        return df2.set_index(["__row__"], drop=False), ["__row__"]

    df2 = df.copy()
    missing = [k for k in keys if k not in df2.columns]
    if missing:
        raise ValueError(f"缺少主键列: {', '.join(missing)}")

    if allow_dup:
        grp = df2.groupby(keys, dropna=False).cumcount()
        df2["__dup__"] = grp.astype(np.int64)
        idx_cols = keys + ["__dup__"]
    else:
        dup = df2.duplicated(subset=keys, keep=False)
        if dup.any():
            sample = df2.loc[dup, keys].head(20).astype(str).to_dict(orient="records")
            raise ValueError(f"主键存在重复（示例前20条）: {sample}")
        idx_cols = keys

    df2 = df2.set_index(idx_cols, drop=False)
    return df2, idx_cols


def _index_to_key_str(idx: pd.Index) -> pd.Series:
    if isinstance(idx, pd.MultiIndex):
        return pd.Series(["|".join("" if v is None or (isinstance(v, float) and math.isnan(v)) else str(v) for v in t) for t in idx.tolist()])
    return pd.Series(["" if v is None or (isinstance(v, float) and math.isnan(v)) else str(v) for v in idx.tolist()])


@dataclass
class SheetDiffResult:
    sheet: str
    left_rows: int
    right_rows: int
    left_only: int
    right_only: int
    common: int
    changed_rows: int
    changed_cells: int
    column_changed_cells: List[Tuple[str, int]]


def diff_sheet(
    left_path: str,
    right_path: str,
    sheet,
    keys: List[str],
    ignore: List[str],
    usecols: Optional[List[str]],
    trim: bool,
    numeric_threshold: float,
    numeric_tol: float,
    allow_dup_keys: bool,
    max_cell_diffs: int,
):
    dl = _read_excel(left_path, sheet=sheet, usecols=usecols)
    dr = _read_excel(right_path, sheet=sheet, usecols=usecols)

    dl.columns = _normalize_columns(dl.columns)
    dr.columns = _normalize_columns(dr.columns)

    left_cols = list(dl.columns)
    right_cols = list(dr.columns)
    if left_cols != right_cols:
        left_set = set(left_cols)
        right_set = set(right_cols)
        missing_in_right = [c for c in left_cols if c not in right_set]
        missing_in_left = [c for c in right_cols if c not in left_set]
        raise ValueError(
            "两文件列不一致: "
            + json.dumps(
                {"missing_in_right": missing_in_right, "missing_in_left": missing_in_left},
                ensure_ascii=False,
            )
        )

    ignore_set = set(ignore)
    compare_cols = [c for c in left_cols if c not in ignore_set and c not in set(keys)]

    dl2, idx_cols = _ensure_unique_index(dl, keys=keys, allow_dup=allow_dup_keys)
    dr2, _ = _ensure_unique_index(dr, keys=keys, allow_dup=allow_dup_keys)

    left_idx = dl2.index
    right_idx = dr2.index

    left_only_idx = left_idx.difference(right_idx)
    right_only_idx = right_idx.difference(left_idx)
    common_idx = left_idx.intersection(right_idx)

    dlc = dl2.loc[common_idx, compare_cols]
    drc = dr2.loc[common_idx, compare_cols]

    diffs = []
    changed_cells = 0
    col_changed = []

    for col in compare_cols:
        a = dlc[col]
        b = drc[col]
        if _is_numeric_like(a, b, threshold=numeric_threshold):
            mask = _numeric_diff_mask(a, b, tol=numeric_tol)
        else:
            mask = _string_diff_mask(a, b, trim=trim)

        n = int(mask.sum())
        if n:
            col_changed.append((col, n))
            changed_cells += n
            if max_cell_diffs > 0:
                idx_hit = common_idx[mask.values]
                if len(idx_hit):
                    df_part = pd.DataFrame(
                        {
                            "key": _index_to_key_str(idx_hit).values,
                            "column": col,
                            "left": a.loc[idx_hit].astype(object).where(~a.loc[idx_hit].isna(), None).astype(str),
                            "right": b.loc[idx_hit].astype(object).where(~b.loc[idx_hit].isna(), None).astype(str),
                        }
                    )
                    diffs.append(df_part)

    diffs_df = pd.concat(diffs, ignore_index=True) if diffs else pd.DataFrame(columns=["key", "column", "left", "right"])
    if max_cell_diffs > 0 and len(diffs_df) > max_cell_diffs:
        diffs_df = diffs_df.head(max_cell_diffs).copy()

    if len(diffs_df):
        changed_rows = int(pd.Series(diffs_df["key"]).nunique())
    else:
        changed_rows = 0

    left_only_df = pd.DataFrame({"key": _index_to_key_str(left_only_idx).values})
    right_only_df = pd.DataFrame({"key": _index_to_key_str(right_only_idx).values})

    changed_rows_df = pd.DataFrame(columns=["key", "changed_columns", "changed_cells"])
    if len(diffs_df):
        g = diffs_df.groupby("key")["column"].agg(list).reset_index()
        g["changed_columns"] = g["column"].map(lambda xs: ",".join(sorted(set(xs))))
        g["changed_cells"] = g["column"].map(len)
        changed_rows_df = g[["key", "changed_columns", "changed_cells"]].sort_values(["changed_cells", "key"], ascending=[False, True])

    result = SheetDiffResult(
        sheet=str(sheet),
        left_rows=int(len(dl2)),
        right_rows=int(len(dr2)),
        left_only=int(len(left_only_idx)),
        right_only=int(len(right_only_idx)),
        common=int(len(common_idx)),
        changed_rows=int(changed_rows),
        changed_cells=int(changed_cells),
        column_changed_cells=sorted(col_changed, key=lambda x: (-x[1], x[0])),
    )

    meta = {
        "sheet": str(sheet),
        "keys": keys if keys else idx_cols,
        "compare_columns": compare_cols,
        "ignore_columns": ignore,
        "index_columns": idx_cols,
    }

    return result, meta, diffs_df, changed_rows_df, left_only_df, right_only_df


def _common_sheets(left_path: str, right_path: str) -> List[str]:
    xl = pd.ExcelFile(left_path)
    xr = pd.ExcelFile(right_path)
    common = [s for s in xl.sheet_names if s in set(xr.sheet_names)]
    if common:
        return common
    return [xl.sheet_names[0]] if xl.sheet_names else [0]


def write_report_xlsx(
    output_path: str,
    all_sheet_results: List[Tuple[SheetDiffResult, dict, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]],
    left_path: str,
    right_path: str,
):
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    summary_rows = []
    for r, meta, *_ in all_sheet_results:
        summary_rows.append(
            {
                "sheet": r.sheet,
                "left_rows": r.left_rows,
                "right_rows": r.right_rows,
                "left_only": r.left_only,
                "right_only": r.right_only,
                "common": r.common,
                "changed_rows": r.changed_rows,
                "changed_cells": r.changed_cells,
            }
        )
    summary_df = pd.DataFrame(summary_rows)

    cols_stat_rows = []
    for r, *_ in all_sheet_results:
        for col, n in r.column_changed_cells:
            cols_stat_rows.append({"sheet": r.sheet, "column": col, "changed_cells": n})
    cols_stat_df = pd.DataFrame(cols_stat_rows).sort_values(["changed_cells", "sheet", "column"], ascending=[False, True, True]) if cols_stat_rows else pd.DataFrame(columns=["sheet", "column", "changed_cells"])

    meta_df = pd.DataFrame(
        [
            {"key": "left_file", "value": str(left_path)},
            {"key": "right_file", "value": str(right_path)},
            {"key": "generated_at", "value": datetime.now().isoformat(timespec="seconds")},
        ]
    )

    with pd.ExcelWriter(str(out), engine="openpyxl") as writer:
        meta_df.to_excel(writer, sheet_name="meta", index=False)
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        cols_stat_df.to_excel(writer, sheet_name="column_stats", index=False)

        for i, (r, meta, diffs_df, changed_rows_df, left_only_df, right_only_df) in enumerate(all_sheet_results, start=1):
            suf = f"_{i}" if i > 1 else ""
            base = _safe_sheet_name(r.sheet)
            meta_s = _safe_sheet_name(f"{base}_meta", suffix=suf)
            diffs_s = _safe_sheet_name(f"{base}_cell_diffs", suffix=suf)
            rows_s = _safe_sheet_name(f"{base}_row_diffs", suffix=suf)
            lo_s = _safe_sheet_name(f"{base}_left_only", suffix=suf)
            ro_s = _safe_sheet_name(f"{base}_right_only", suffix=suf)

            pd.DataFrame(
                [{"key": k, "value": json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict)) else str(v)} for k, v in meta.items()]
            ).to_excel(writer, sheet_name=meta_s, index=False)
            diffs_df.to_excel(writer, sheet_name=diffs_s, index=False)
            changed_rows_df.to_excel(writer, sheet_name=rows_s, index=False)
            left_only_df.to_excel(writer, sheet_name=lo_s, index=False)
            right_only_df.to_excel(writer, sheet_name=ro_s, index=False)

        wb = writer.book
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.max_row <= 1 or ws.max_column <= 1:
                continue
            df_preview = pd.DataFrame(ws.values)
            df_preview.columns = df_preview.iloc[0].astype(str).tolist()
            df_preview = df_preview.iloc[1:]
            _auto_columns_width(ws, df_preview)


def write_report_html(
    output_path: str,
    all_sheet_results: List[Tuple[SheetDiffResult, dict, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]],
    left_path: str,
    right_path: str,
    max_keys_sample: int,
    max_row_diffs: int,
    max_col_stats: int,
):
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    report: Dict[str, object] = {
        "meta": {
            "left_file": str(left_path),
            "right_file": str(right_path),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        },
        "summary": [],
        "sheets": [],
    }

    for r, meta, diffs_df, changed_rows_df, left_only_df, right_only_df in all_sheet_results:
        report["summary"].append(
            {
                "sheet": r.sheet,
                "left_rows": r.left_rows,
                "right_rows": r.right_rows,
                "left_only": r.left_only,
                "right_only": r.right_only,
                "common": r.common,
                "changed_rows": r.changed_rows,
                "changed_cells": r.changed_cells,
            }
        )

        col_stats = [{"column": c, "changed_cells": n} for c, n in r.column_changed_cells[: max(0, max_col_stats)]]
        row_diffs_show = changed_rows_df.head(max(0, max_row_diffs)).to_dict(orient="records")
        cell_diffs_show = diffs_df.to_dict(orient="records")
        left_only_sample = left_only_df.head(max(0, max_keys_sample))["key"].astype(str).tolist() if len(left_only_df) else []
        right_only_sample = right_only_df.head(max(0, max_keys_sample))["key"].astype(str).tolist() if len(right_only_df) else []

        report["sheets"].append(
            {
                "sheet": r.sheet,
                "metrics": {
                    "left_rows": r.left_rows,
                    "right_rows": r.right_rows,
                    "left_only": r.left_only,
                    "right_only": r.right_only,
                    "common": r.common,
                    "changed_rows": r.changed_rows,
                    "changed_cells": r.changed_cells,
                    "shown_row_diffs": int(min(len(changed_rows_df), max(0, max_row_diffs))),
                    "shown_cell_diffs": int(len(diffs_df)),
                    "shown_left_only_keys": int(min(len(left_only_df), max(0, max_keys_sample))),
                    "shown_right_only_keys": int(min(len(right_only_df), max(0, max_keys_sample))),
                },
                "config": meta,
                "column_stats": col_stats,
                "row_diffs": row_diffs_show,
                "cell_diffs": cell_diffs_show,
                "left_only_keys": left_only_sample,
                "right_only_keys": right_only_sample,
            }
        )

    data_json = json.dumps(report, ensure_ascii=False)
    data_json_safe = data_json.replace("</", "<\\/")
    summary_rows = []
    for s in report["summary"]:
        summary_rows.append(
            f"<tr>"
            f"<td>{_escape_html(s['sheet'])}</td>"
            f"<td>{s['left_rows']}</td>"
            f"<td>{s['right_rows']}</td>"
            f"<td>{s['left_only']}</td>"
            f"<td>{s['right_only']}</td>"
            f"<td>{s['changed_rows']}</td>"
            f"<td>{s['changed_cells']}</td>"
            f"</tr>"
        )
    static_summary_html = (
        "<section class=\"wrap\" style=\"padding: 12px 18px 0 18px;\">"
        "<div class=\"card\" style=\"border-radius:12px; padding:12px;\">"
        "<div class=\"k\" style=\"margin-bottom:8px;\">静态摘要（若下方未渲染，此处仍可用）</div>"
        "<div style=\"overflow:auto;\">"
        "<table>"
        "<thead><tr>"
        "<th>sheet</th><th>left_rows</th><th>right_rows</th><th>left_only</th><th>right_only</th>"
        "<th>changed_rows</th><th>changed_cells</th>"
        "</tr></thead>"
        "<tbody>"
        + "".join(summary_rows)
        + "</tbody></table></div></div></section>"
    )

    template = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Excel Diff Report</title>
  <style>
    :root {{
      --bg: #0b1220;
      --panel: #0f1b31;
      --panel2: #111f3a;
      --text: #e7eefc;
      --muted: #9bb0d4;
      --border: rgba(255,255,255,0.10);
      --ok: #2dd4bf;
      --warn: #f59e0b;
      --bad: #fb7185;
      --mono: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
      --sans: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial, "Apple Color Emoji", "Segoe UI Emoji";
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: var(--sans);
      background: radial-gradient(1200px 600px at 10% 0%, rgba(45,212,191,0.12), transparent 55%),
                  radial-gradient(1000px 500px at 90% 10%, rgba(251,113,133,0.10), transparent 55%),
                  var(--bg);
      color: var(--text);
    }}
    header {{
      padding: 18px 18px 10px 18px;
      border-bottom: 1px solid var(--border);
      background: linear-gradient(180deg, rgba(255,255,255,0.04), transparent);
    }}
    .wrap {{ max-width: 1200px; margin: 0 auto; }}
    h1 {{ font-size: 18px; margin: 0 0 8px 0; }}
    .meta {{ color: var(--muted); font-size: 12px; line-height: 1.4; }}
    .meta code {{ font-family: var(--mono); color: var(--text); }}
    .toolbar {{
      display: flex; gap: 10px; flex-wrap: wrap;
      padding: 12px 18px; border-bottom: 1px solid var(--border);
      background: rgba(255,255,255,0.02);
    }}
    .tabs {{ display: flex; gap: 8px; flex-wrap: wrap; }}
    button.tab {{
      border: 1px solid var(--border);
      background: rgba(255,255,255,0.03);
      color: var(--text);
      padding: 6px 10px;
      border-radius: 10px;
      cursor: pointer;
      font-size: 12px;
    }}
    button.tab.active {{
      border-color: rgba(45,212,191,0.55);
      background: rgba(45,212,191,0.12);
    }}
    main {{ padding: 18px; }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 10px;
      margin-bottom: 14px;
    }}
    .card {{
      border: 1px solid var(--border);
      background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.015));
      border-radius: 14px;
      padding: 10px 12px;
      min-height: 78px;
    }}
    .k {{ color: var(--muted); font-size: 12px; margin-bottom: 6px; }}
    .v {{ font-family: var(--mono); font-size: 18px; }}
    .v.ok {{ color: var(--ok); }}
    .v.bad {{ color: var(--bad); }}
    .v.warn {{ color: var(--warn); }}
    section {{
      border: 1px solid var(--border);
      background: rgba(255,255,255,0.02);
      border-radius: 14px;
      padding: 12px;
      margin: 12px 0;
    }}
    section h2 {{ margin: 0 0 10px 0; font-size: 14px; }}
    .row {{
      display: flex; gap: 12px; flex-wrap: wrap; align-items: center;
      margin-bottom: 10px;
    }}
    input[type="text"] {{
      width: min(520px, 100%);
      border-radius: 10px;
      border: 1px solid var(--border);
      background: rgba(255,255,255,0.03);
      color: var(--text);
      padding: 8px 10px;
      font-family: var(--mono);
      font-size: 12px;
      outline: none;
    }}
    .hint {{ color: var(--muted); font-size: 12px; }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 12px;
    }}
    th, td {{
      border-bottom: 1px solid rgba(255,255,255,0.06);
      padding: 7px 8px;
      vertical-align: top;
    }}
    th {{
      position: sticky;
      top: 0;
      background: rgba(15,27,49,0.96);
      z-index: 1;
      text-align: left;
      color: var(--muted);
      font-weight: 600;
    }}
    td code {{ font-family: var(--mono); }}
    .mono {{ font-family: var(--mono); }}
    details > summary {{
      cursor: pointer;
      color: var(--muted);
      font-size: 12px;
    }}
    .hidden {{ display: none; }}
    @media (max-width: 980px) {{
      .grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
    }}
    @media (max-width: 520px) {{
      .grid {{ grid-template-columns: repeat(1, minmax(0, 1fr)); }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="wrap">
      <h1>Excel 对比报告</h1>
      <div class="meta">
        <div>left: <code>@@LEFT@@</code></div>
        <div>right: <code>@@RIGHT@@</code></div>
        <div>generated_at: <code>@@GENERATED_AT@@</code></div>
      </div>
    </div>
  </header>
  <div class="toolbar">
    <div class="tabs" id="tabs"></div>
  </div>
  @@STATIC_SUMMARY@@
  <main class="wrap">
    <div id="content"></div>
  </main>
  <script id="report-data" type="application/json">@@DATA_JSON@@</script>
  <script>
    const data = JSON.parse(document.getElementById('report-data').textContent);
    const tabsEl = document.getElementById('tabs');
    const contentEl = document.getElementById('content');

    function fmt(n) {{
      if (n === null || n === undefined) return '';
      const x = Number(n);
      if (Number.isNaN(x)) return String(n);
      return x.toLocaleString('zh-CN');
    }}

    function trendClass(changedCells, leftOnly, rightOnly) {{
      if ((leftOnly + rightOnly) > 0) return 'warn';
      if (changedCells > 0) return 'bad';
      return 'ok';
    }}

    function mkEl(tag, attrs={{}}, text=null) {{
      const el = document.createElement(tag);
      for (const [k,v] of Object.entries(attrs)) {{
        if (k === 'class') el.className = v;
        else el.setAttribute(k, v);
      }}
      if (text !== null) el.textContent = text;
      return el;
    }}

    function renderTable(container, cols, rows, rowLimit=null) {{
      const table = mkEl('table');
      const thead = mkEl('thead');
      const trh = mkEl('tr');
      for (const c of cols) trh.appendChild(mkEl('th', {{}}, c));
      thead.appendChild(trh);
      table.appendChild(thead);
      const tbody = mkEl('tbody');
      const n = rowLimit ? Math.min(rows.length, rowLimit) : rows.length;
      for (let i=0; i<n; i++) {{
        const r = rows[i];
        const tr = mkEl('tr');
        for (const c of cols) {{
          const td = mkEl('td');
          const val = r[c];
          td.textContent = (val === null || val === undefined) ? '' : String(val);
          tr.appendChild(td);
        }}
        tbody.appendChild(tr);
      }}
      table.appendChild(tbody);
      container.appendChild(table);
      return table;
    }}

    function renderSheet(sheetData) {{
      const m = sheetData.metrics;
      const root = mkEl('div', {{}});

      const grid = mkEl('div', {{class: 'grid'}});
      const items = [
        ['行数(left)', fmt(m.left_rows), 'mono'],
        ['行数(right)', fmt(m.right_rows), 'mono'],
        ['仅在 left', fmt(m.left_only), `v ${trendClass(0, m.left_only, 0)}`],
        ['仅在 right', fmt(m.right_only), `v ${trendClass(0, 0, m.right_only)}`],
        ['共同 key', fmt(m.common), 'mono'],
        ['变更行数', fmt(m.changed_rows), `v ${m.changed_rows ? 'bad' : 'ok'}`],
        ['变更单元格', fmt(m.changed_cells), `v ${m.changed_cells ? 'bad' : 'ok'}`],
        ['输出 cell 差异', fmt(m.shown_cell_diffs), 'mono']
      ];

      for (const [k, v, vClass] of items) {{
        const card = mkEl('div', {{class:'card'}});
        card.appendChild(mkEl('div', {{class:'k'}}, k));
        card.appendChild(mkEl('div', {{class: (vClass.startsWith('v') ? vClass : 'v '+vClass)}}, v));
        grid.appendChild(card);
      }}
      root.appendChild(grid);

      const secCols = mkEl('section');
      secCols.appendChild(mkEl('h2', {{}}, '差异列统计（Top）'));
      if (!sheetData.column_stats || sheetData.column_stats.length === 0) {{
        secCols.appendChild(mkEl('div', {{class:'hint'}}, '无差异列'));
      }} else {{
        renderTable(secCols, ['column','changed_cells'], sheetData.column_stats);
      }}
      root.appendChild(secCols);

      const secRows = mkEl('section');
      secRows.appendChild(mkEl('h2', {{}}, '变更行汇总（Top）'));
      secRows.appendChild(mkEl('div', {{class:'hint'}}, `仅展示前 ${fmt(m.shown_row_diffs)} 条（按 changed_cells 降序）`));
      if (!sheetData.row_diffs || sheetData.row_diffs.length === 0) {{
        secRows.appendChild(mkEl('div', {{class:'hint'}}, '无变更行'));
      }} else {{
        renderTable(secRows, ['key','changed_cells','changed_columns'], sheetData.row_diffs);
      }}
      root.appendChild(secRows);

      const secCell = mkEl('section');
      secCell.appendChild(mkEl('h2', {{}}, '单元格差异（可搜索）'));
      const row = mkEl('div', {{class:'row'}});
      const inp = mkEl('input', {{type:'text', placeholder:'搜索 key/column/left/right（本页内过滤）'}});
      row.appendChild(inp);
      row.appendChild(mkEl('div', {{class:'hint'}}, `展示 ${fmt(m.shown_cell_diffs)} 条（如 changed_cells 更大表示已截断）`));
      secCell.appendChild(row);
      const tableWrap = mkEl('div', {{}});
      secCell.appendChild(tableWrap);
      if (!sheetData.cell_diffs || sheetData.cell_diffs.length === 0) {{
        tableWrap.appendChild(mkEl('div', {{class:'hint'}}, '未输出 cell 差异（可能设置了 max-cell-diffs=0 或无差异）'));
      }} else {{
        const rows = sheetData.cell_diffs.map(r => {{
          return {{
            key: r.key ?? '',
            column: r.column ?? '',
            left: r.left ?? '',
            right: r.right ?? ''
          }};
        }});
        const cols = ['key','column','left','right'];
        const table = renderTable(tableWrap, cols, rows);
        inp.addEventListener('input', () => {{
          const q = inp.value.trim().toLowerCase();
          const trs = table.querySelectorAll('tbody tr');
          trs.forEach(tr => {{
            const t = tr.textContent.toLowerCase();
            tr.style.display = q === '' || t.includes(q) ? '' : 'none';
          }});
        }});
      }}
      root.appendChild(secCell);

      const secOnly = mkEl('section');
      secOnly.appendChild(mkEl('h2', {{}}, '仅存在于一侧的 key（抽样）'));
      const d1 = mkEl('details');
      d1.appendChild(mkEl('summary', {{}}, `left_only keys 抽样：显示 ${fmt(m.shown_left_only_keys)} / ${fmt(m.left_only)}`));
      const pre1 = mkEl('pre', {{class:'mono'}}, (sheetData.left_only_keys || []).join('\\n'));
      d1.appendChild(pre1);
      secOnly.appendChild(d1);
      const d2 = mkEl('details');
      d2.appendChild(mkEl('summary', {{}}, `right_only keys 抽样：显示 ${fmt(m.shown_right_only_keys)} / ${fmt(m.right_only)}`));
      const pre2 = mkEl('pre', {{class:'mono'}}, (sheetData.right_only_keys || []).join('\\n'));
      d2.appendChild(pre2);
      secOnly.appendChild(d2);

      const d3 = mkEl('details');
      d3.appendChild(mkEl('summary', {{}}, '对比配置'));
      const pre3 = mkEl('pre', {{class:'mono'}}, JSON.stringify(sheetData.config, null, 2));
      d3.appendChild(pre3);
      secOnly.appendChild(d3);

      root.appendChild(secOnly);
      return root;
    }}

    function setActive(sheetName) {{
      [...tabsEl.querySelectorAll('button.tab')].forEach(b => {{
        b.classList.toggle('active', b.dataset.sheet === sheetName);
      }});
      contentEl.innerHTML = '';
      const sheet = data.sheets.find(s => s.sheet === sheetName);
      if (!sheet) {{
        contentEl.appendChild(mkEl('div', {{}}, '未找到 sheet'));
        return;
      }}
      contentEl.appendChild(renderSheet(sheet));
      window.scrollTo({{ top: 0, behavior: 'smooth' }});
    }}

    function init() {{
      for (const s of data.sheets) {{
        const b = mkEl('button', {{class:'tab'}}, s.sheet);
        b.dataset.sheet = s.sheet;
        b.addEventListener('click', () => setActive(s.sheet));
        tabsEl.appendChild(b);
      }}
      if (data.sheets.length) setActive(data.sheets[0].sheet);
    }}
    init();
  </script>
</body>
</html>
"""

    html = (
        template.replace("@@LEFT@@", _escape_html(left_path))
        .replace("@@RIGHT@@", _escape_html(right_path))
        .replace("@@GENERATED_AT@@", _escape_html(report["meta"]["generated_at"]))
        .replace("@@STATIC_SUMMARY@@", static_summary_html)
        .replace("@@DATA_JSON@@", data_json_safe)
    )

    out.write_text(html, encoding="utf-8")


def main(argv: Optional[Sequence[str]] = None) -> int:
    p = argparse.ArgumentParser(description="对比两个结构相同的 Excel（支持按主键对齐、输出差异报告）")
    p.add_argument("--left", required=True, help="左侧 Excel 文件路径")
    p.add_argument("--right", required=True, help="右侧 Excel 文件路径")
    p.add_argument("--sheet", default="all", help='sheet 名称/序号(从0开始)/all（默认 all）')
    p.add_argument("--keys", default="", help="主键列名，逗号分隔（推荐；为空则按行号对齐）")
    p.add_argument("--ignore", default="", help="忽略列名，逗号分隔")
    p.add_argument("--usecols", default="", help="仅读取指定列（逗号分隔；为空则读取全部列）")
    p.add_argument("--trim", action="store_true", help="对字符串列做 strip 后再比较")
    p.add_argument("--numeric-threshold", type=float, default=0.9, help="把一列判定为数值列的非空数值占比阈值（默认 0.9）")
    p.add_argument("--numeric-tol", type=float, default=0.0, help="数值列比较容差（默认 0）")
    p.add_argument("--allow-dup-keys", action="store_true", help="允许主键重复（将按出现顺序追加 __dup__ 对齐）")
    p.add_argument("--max-cell-diffs", type=int, default=200000, help="cell 级差异最大输出条数（默认 200000；0 表示不输出 cell 差异）")
    p.add_argument("--max-row-diffs", type=int, default=5000, help="变更行汇总最大输出条数（默认 5000）")
    p.add_argument("--max-col-stats", type=int, default=60, help="差异列统计最大输出条数（默认 60）")
    p.add_argument("--max-keys-sample", type=int, default=500, help="left_only/right_only key 抽样最大行数（默认 500）")
    p.add_argument("--output", default="", help="输出报告路径（.html 或 .xlsx；默认自动生成 .html）")

    args = p.parse_args(argv)

    left_path = str(Path(args.left).expanduser())
    right_path = str(Path(args.right).expanduser())
    sheet_arg = _coerce_sheet_arg(args.sheet)
    keys = _to_list(args.keys)
    ignore = _to_list(args.ignore)
    usecols = _to_list(args.usecols)
    usecols = usecols if usecols else None

    if not args.output:
        lb = Path(left_path).stem
        rb = Path(right_path).stem
        tag = _now_tag()
        args.output = str(Path.cwd() / f"excel_diff__{lb}__vs__{rb}__{tag}.html")

    if sheet_arg == "all":
        sheets = _common_sheets(left_path, right_path)
    else:
        sheets = [sheet_arg]

    all_results = []
    for s in sheets:
        r, meta, diffs_df, changed_rows_df, left_only_df, right_only_df = diff_sheet(
            left_path=left_path,
            right_path=right_path,
            sheet=s,
            keys=keys,
            ignore=ignore,
            usecols=usecols,
            trim=bool(args.trim),
            numeric_threshold=float(args.numeric_threshold),
            numeric_tol=float(args.numeric_tol),
            allow_dup_keys=bool(args.allow_dup_keys),
            max_cell_diffs=int(args.max_cell_diffs),
        )
        all_results.append((r, meta, diffs_df, changed_rows_df, left_only_df, right_only_df))

    out_path = str(args.output)
    if out_path.lower().endswith(".xlsx"):
        write_report_xlsx(out_path, all_results, left_path=left_path, right_path=right_path)
    else:
        write_report_html(
            out_path,
            all_results,
            left_path=left_path,
            right_path=right_path,
            max_keys_sample=int(args.max_keys_sample),
            max_row_diffs=int(args.max_row_diffs),
            max_col_stats=int(args.max_col_stats),
        )

    out = {
        "output": str(args.output),
        "sheets": [
            {
                "sheet": r.sheet,
                "left_rows": r.left_rows,
                "right_rows": r.right_rows,
                "left_only": r.left_only,
                "right_only": r.right_only,
                "common": r.common,
                "changed_rows": r.changed_rows,
                "changed_cells": r.changed_cells,
            }
            for (r, *_rest) in all_results
        ],
    }
    sys.stdout.write(json.dumps(out, ensure_ascii=False, indent=2) + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
