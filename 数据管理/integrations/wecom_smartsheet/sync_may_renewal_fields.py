#!/usr/bin/env python3
"""Refresh five fields in the May renewal tracking WeCom smart sheet.

This script intentionally updates existing rows only. A WeCom smartsheet
webhook cannot read rows, so the first run must prime a local VIN -> record_id
state from wecom-cli or from an exported get_records JSON response. Daily runs
then use the state file and update records by record_id.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import subprocess
import sys
import time as time_mod
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, Iterable

import duckdb
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
DATA_ROOT = HERE.parents[1]
REPO_ROOT = HERE.parents[2]
DEFAULT_STATE_PATH = HERE / "state" / "renewal_2026_may_jul_vin_record_map_6fDmy0.json"
DEFAULT_TABLE_SCHEMA_PATH = HERE / "outputs" / "renewal_may_jul_schema_6fDmy0.json"
DEFAULT_POLICY_GLOB = DATA_ROOT / "warehouse" / "fact" / "policy" / "current" / "*.parquet"
DEFAULT_RENEWAL_TRACKER_PATH = DATA_ROOT / "warehouse" / "fact" / "renewal_tracker" / "latest.parquet"
DEFAULT_QUOTES_PATH = DATA_ROOT / "warehouse" / "fact" / "quotes_conversion" / "latest.parquet"
DEFAULT_CUSTOMER_FLOW_PATH = DATA_ROOT / "warehouse" / "fact" / "customer_flow" / "latest.parquet"

VIN_FIELD_ID = "fcCW6A"
FIELD_IDS = {
    "is_renewed": "fwuflw",       # 是否成交
    "is_quoted": "fyIIAm",        # 是否报价
    "insurance_grade": "fq15NT",  # 风险等级
    "pricing_factor": "fvCvWp",   # 自主系数
    "loss_company": "fXXEDk",     # 流失公司
}
FULL_FIELD_IDS = {
    "list_type": "f3ShQB",
    "insurance_combo": "fCUIIa",
    "plate_no": "fPJzr7",
    "vehicle_frame_no": VIN_FIELD_ID,
    "expiry_date": "fd0gUV",
    "owner": "feXYXJ",
    "owner_user": "f54Wcl",
    "seat_account": "fgKOrG",
    "team": "fjg2u5",
    "vehicle_type": "fqrd28",
    "coverage_combination": "fQrZnC",
    **FIELD_IDS,
}
KEY_LABELS = {
    "list_type": "名单类型",
    "insurance_combo": "投保险种",
    "plate_no": "车牌号",
    "vehicle_frame_no": "车架号",
    "expiry_date": "到期日",
    "owner": "归属人",
    "owner_user": "归属人(人员)",
    "seat_account": "坐席域账号",
    "team": "归属团队",
    "vehicle_type": "车型",
    "coverage_combination": "险别组合",
    "is_renewed": "是否成交",
    "is_quoted": "是否报价",
    "insurance_grade": "风险等级",
    "pricing_factor": "自主系数",
    "loss_company": "流失公司",
}
LABEL_TO_KEY = {
    **{label: key for key, label in KEY_LABELS.items()},
    "保单到期时间": "expiry_date",
}
DEFAULT_FIELD_TYPES = {
    "list_type": "select",
    "insurance_combo": "select",
    "plate_no": "text",
    "vehicle_frame_no": "text",
    "expiry_date": "date",
    "owner": "text",
    "owner_user": "user",
    "seat_account": "text",
    "team": "select",
    "vehicle_type": "select",
    "coverage_combination": "select",
    "is_renewed": "select",
    "is_quoted": "select",
    "insurance_grade": "select",
    "pricing_factor": "number",
    "loss_company": "text",
}
CURRENT_FULL_FIELD_IDS = dict(FULL_FIELD_IDS)
CURRENT_FIELD_TYPES = dict(DEFAULT_FIELD_TYPES)
DEFAULT_WEBHOOK_ENV = "WECOM_SMARTSHEET_WEBHOOK_RENEWAL_MAY"
FIELD_SETS = {
    # 日常默认只写报价相关稳定字段；风险等级/流失公司先观察，避免目标表单选项不全时阻塞报价更新。
    "quote": ["is_renewed", "is_quoted", "pricing_factor"],
    "all": ["is_renewed", "is_quoted", "insurance_grade", "pricing_factor", "loss_company"],
}
SEED_UPDATE_KEYS = ["is_renewed", "is_quoted", "coverage_combination", "insurance_grade", "pricing_factor", "loss_company"]
BASE_UPDATE_KEYS = list(FIELD_SETS["quote"])
BASE_SEED_KEYS = list(FULL_FIELD_IDS.keys())


@dataclass(frozen=True)
class SyncConfig:
    start: str
    end: str
    quote_window_start: str
    renewal_tracker_path: str
    quotes_path: str
    customer_flow_path: str
    policy_glob: str
    state_path: Path
    webhook_env: str
    webhook_url: str | None
    batch_size: int
    force: bool
    vin_filter: set[str] | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="更新 5 月续保跟踪表的是否成交/是否报价/风险等级/自主系数/流失公司字段"
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    def add_common(p: argparse.ArgumentParser) -> None:
        p.add_argument("--state", default=str(DEFAULT_STATE_PATH), help="VIN -> record_id state 路径")

    prime = subparsers.add_parser("prime-state", help="从企微现有表读取车架号并建立 record_id state")
    add_common(prime)
    src = prime.add_mutually_exclusive_group(required=True)
    src.add_argument("--records-json", help="wecom-cli smartsheet_get_records 的 JSON 输出文件")
    src.add_argument("--docid", help="智能表 docid；需同时传 --sheet-id")
    src.add_argument("--url", help="智能表 URL；需同时传 --sheet-id")
    prime.add_argument("--sheet-id", help="智能表 sheet_id")
    prime.add_argument("--vin-field-id", default=VIN_FIELD_ID)
    prime.add_argument("--vin-field-title", default="车架号")

    sync = subparsers.add_parser("sync", help="按本地数据源更新 5 个字段；默认 dry-run")
    add_common(sync)
    sync.add_argument("--start", default="2026-05-01", help="到期日开始，默认 2026-05-01")
    sync.add_argument("--end", default="2026-07-31", help="到期日结束，默认 2026-07-31")
    sync.add_argument("--quote-window-start", default="2025-12-03", help="报价匹配开始日期")
    sync.add_argument("--renewal-tracker-path", default=str(DEFAULT_RENEWAL_TRACKER_PATH))
    sync.add_argument("--quotes-path", default=str(DEFAULT_QUOTES_PATH))
    sync.add_argument("--customer-flow-path", default=str(DEFAULT_CUSTOMER_FLOW_PATH))
    sync.add_argument("--policy-glob", default=str(DEFAULT_POLICY_GLOB))
    sync.add_argument("--webhook-env", default=DEFAULT_WEBHOOK_ENV)
    sync.add_argument("--webhook-url", help="直接传 webhook URL；优先级高于 --webhook-env，不建议写入脚本")
    sync.add_argument(
        "--table-schema-file",
        help="目标表 webhook 示例 JSON，含 schema/add_records，用于覆盖字段 ID 和字段类型",
    )
    sync.add_argument("--batch-size", type=int, default=100)
    sync.add_argument(
        "--fields",
        choices=sorted(FIELD_SETS),
        default="quote",
        help="同步字段集：quote=是否成交/是否报价/自主系数；all=再包含风险等级/流失公司",
    )
    sync.add_argument("--force", action="store_true", help="即使 payload hash 未变化也更新")
    sync.add_argument("--execute", action="store_true", help="真实调用 webhook；默认只 dry-run")
    sync.add_argument("--vin", action="append", default=[], help="只更新指定车架号；可多次传，也可用英文逗号分隔")
    sync.add_argument("--vin-file", help="只更新文件中的车架号；每行一个，支持逗号分隔")

    inspect = subparsers.add_parser("inspect", help="只查询指定车架号的 5 个字段，不读/写企微")
    inspect.add_argument("--start", default="2026-05-01", help="到期日开始，默认 2026-05-01")
    inspect.add_argument("--end", default="2026-07-31", help="到期日结束，默认 2026-07-31")
    inspect.add_argument("--quote-window-start", default="2025-12-03", help="报价匹配开始日期")
    inspect.add_argument("--renewal-tracker-path", default=str(DEFAULT_RENEWAL_TRACKER_PATH))
    inspect.add_argument("--quotes-path", default=str(DEFAULT_QUOTES_PATH))
    inspect.add_argument("--customer-flow-path", default=str(DEFAULT_CUSTOMER_FLOW_PATH))
    inspect.add_argument("--policy-glob", default=str(DEFAULT_POLICY_GLOB))
    inspect.add_argument("--vin", action="append", default=[], help="指定车架号；可多次传，也可用英文逗号分隔")
    inspect.add_argument("--vin-file", help="车架号文件；每行一个，支持逗号分隔")
    inspect.add_argument("--output", help="把核验清单写入 JSON 文件")

    seed = subparsers.add_parser("seed-from-excel", help="从导出的 Excel 三个月清单全字段写入空表/新表，并保存 record_id state")
    seed.add_argument("--input", required=True, help="企微智能表导出的 Excel")
    seed.add_argument("--start", default="2026-05-01", help="到期日开始，默认 2026-05-01")
    seed.add_argument("--end", default="2026-07-31", help="到期日结束，默认 2026-07-31")
    seed.add_argument("--quote-window-start", default="2025-12-03", help="报价匹配开始日期")
    seed.add_argument("--renewal-tracker-path", default=str(DEFAULT_RENEWAL_TRACKER_PATH))
    seed.add_argument("--quotes-path", default=str(DEFAULT_QUOTES_PATH))
    seed.add_argument("--customer-flow-path", default=str(DEFAULT_CUSTOMER_FLOW_PATH))
    seed.add_argument("--policy-glob", default=str(DEFAULT_POLICY_GLOB))
    seed.add_argument("--state", default=str(DEFAULT_STATE_PATH), help="VIN -> record_id state 路径")
    seed.add_argument("--webhook-env", default=DEFAULT_WEBHOOK_ENV)
    seed.add_argument("--webhook-url", help="直接传 webhook URL；优先级高于 --webhook-env")
    seed.add_argument(
        "--table-schema-file",
        help="目标表 webhook 示例 JSON，含 schema/add_records，用于覆盖字段 ID 和字段类型",
    )
    seed.add_argument("--batch-size", type=int, default=100)
    seed.add_argument("--execute", action="store_true", help="真实调用 webhook；默认只 dry-run")
    seed.add_argument("--allow-existing-state", action="store_true", help="允许 state 已存在记录时继续 seed（默认拒绝，避免重复新增）")
    seed.add_argument("--output", help="把 dry-run 计划写入 JSON 文件")
    return parser.parse_args()


def load_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"version": 1, "vin_field_id": VIN_FIELD_ID, "records": {}}
    return json.loads(path.read_text(encoding="utf-8"))


def load_dotenv_local(path: Path = REPO_ROOT / ".env.local") -> None:
    """Lightweight .env.local loader; keeps existing process env authoritative."""
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'").strip('"')
        if key and key not in os.environ:
            os.environ[key] = value


def save_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def extract_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return str(value).strip()
    if isinstance(value, list):
        parts: list[str] = []
        for item in value:
            if isinstance(item, dict):
                parts.append(str(item.get("text") or item.get("value") or "").strip())
            else:
                parts.append(extract_text(item))
        return "".join(parts).strip()
    if isinstance(value, dict):
        return str(value.get("text") or value.get("value") or "").strip()
    return str(value).strip()


def load_records_json(path: Path) -> dict[str, Any]:
    raw = path.read_text(encoding="utf-8")
    return json.loads(raw)


def normalize_wecom_response(response: dict[str, Any]) -> dict[str, Any]:
    """Unwrap wecom-cli MCP-style output and fail on non-zero WeCom errcode."""
    payload = response
    result = response.get("result")
    if isinstance(result, dict) and isinstance(result.get("content"), list):
        for item in result["content"]:
            if not isinstance(item, dict) or "text" not in item:
                continue
            text = str(item["text"])
            try:
                payload = json.loads(text)
                break
            except json.JSONDecodeError:
                continue

    errcode = payload.get("errcode")
    if errcode not in (None, 0):
        help_message = payload.get("help_message")
        detail = help_message or payload.get("errmsg") or payload
        raise RuntimeError(f"企业微信读取记录失败 errcode={errcode}: {detail}")
    return payload


def get_records_with_wecom_cli(*, docid: str | None, url: str | None, sheet_id: str, vin_field_id: str) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "sheet_id": sheet_id,
    }
    if docid:
        payload["docid"] = docid
    elif url:
        payload["url"] = url
    else:
        raise ValueError("docid/url 至少需要一个")

    proc = subprocess.run(
        ["wecom-cli", "doc", "smartsheet_get_records", json.dumps(payload, ensure_ascii=False)],
        check=False,
        capture_output=True,
        text=True,
    )
    if proc.returncode != 0:
        raise RuntimeError(f"wecom-cli smartsheet_get_records 失败: {proc.stderr.strip()}")
    return normalize_wecom_response(json.loads(proc.stdout))


def iter_records(response: dict[str, Any]) -> Iterable[dict[str, Any]]:
    if isinstance(response.get("records"), list):
        yield from response["records"]
        return
    data = response.get("data")
    if isinstance(data, dict) and isinstance(data.get("records"), list):
        yield from data["records"]
        return
    if isinstance(response.get("items"), list):
        yield from response["items"]


def prime_state_from_records(
    response: dict[str, Any],
    *,
    state_path: Path,
    vin_field_id: str,
    vin_field_title: str,
) -> dict[str, Any]:
    response = normalize_wecom_response(response)
    state = load_state(state_path)
    records = state.setdefault("records", {})
    duplicates: dict[str, int] = {}
    added = 0
    missing_vin = 0
    missing_record_id = 0

    for rec in iter_records(response):
        rid = rec.get("record_id") or rec.get("id")
        if not rid:
            missing_record_id += 1
            continue
        values = rec.get("values") or {}
        vin = extract_text(values.get(vin_field_id)) or extract_text(values.get(vin_field_title))
        if not vin:
            missing_vin += 1
            continue
        if vin in records:
            duplicates[vin] = duplicates.get(vin, 1) + 1
        records[vin] = {
            **records.get(vin, {}),
            "record_id": rid,
            "primed_at": datetime.now(timezone.utc).isoformat(),
        }
        added += 1

    state["vin_field_id"] = vin_field_id
    state["summary"] = {
        "operation": "prime-state",
        "records_after": len(records),
        "records_seen": added,
        "missing_vin": missing_vin,
        "missing_record_id": missing_record_id,
        "duplicate_vin_count": len(duplicates),
        "duplicate_vin_sample": sorted(duplicates)[:20],
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    save_state(state_path, state)
    return {"state_path": str(state_path), **state["summary"]}


def fetch_source_rows(config: SyncConfig) -> list[dict[str, Any]]:
    sql = f"""
    WITH rt AS (
      SELECT * EXCLUDE rn FROM (
        SELECT *,
               ROW_NUMBER() OVER (
                 PARTITION BY vehicle_frame_no
                 ORDER BY expiry_date DESC NULLS LAST, source_policy_no DESC
               ) AS rn
        FROM read_parquet('{config.renewal_tracker_path}')
        WHERE vehicle_frame_no IS NOT NULL AND vehicle_frame_no != ''
          AND CAST(expiry_date AS DATE) BETWEEN CAST(? AS DATE) AND CAST(? AS DATE)
      ) WHERE rn = 1
    ),
    q_latest AS (
      SELECT * EXCLUDE rn FROM (
        SELECT vehicle_frame_no,
               insurance_grade AS quote_insurance_grade,
               commercial_pricing_factor AS quote_pricing_factor,
               quote_time,
               ROW_NUMBER() OVER (
                 PARTITION BY vehicle_frame_no
                 ORDER BY quote_time DESC NULLS LAST
               ) AS rn
        FROM read_parquet('{config.quotes_path}')
        WHERE vehicle_frame_no IS NOT NULL AND vehicle_frame_no != ''
          AND insurance_type = '商业保险'
          AND CAST(quote_time AS DATE) >= CAST(? AS DATE)
      ) WHERE rn = 1
    ),
    policy_one AS (
      SELECT policy_no,
             vehicle_frame_no,
             ANY_VALUE(insurance_grade) FILTER (
               WHERE insurance_grade IS NOT NULL AND insurance_grade != ''
             ) AS policy_insurance_grade,
             ANY_VALUE(commercial_pricing_factor) FILTER (
               WHERE commercial_pricing_factor IS NOT NULL
             ) AS policy_pricing_factor
      FROM read_parquet('{config.policy_glob}', union_by_name=true)
      WHERE vehicle_frame_no IS NOT NULL AND vehicle_frame_no != ''
      GROUP BY policy_no, vehicle_frame_no
    ),
    flow AS (
      SELECT vehicle_frame_no,
             ANY_VALUE(NULLIF(NULLIF(next_insurer, ''), 'NaN')) FILTER (
               WHERE next_insurer IS NOT NULL AND next_insurer != ''
             ) AS next_insurer
      FROM read_parquet('{config.customer_flow_path}')
      WHERE vehicle_frame_no IS NOT NULL AND vehicle_frame_no != ''
      GROUP BY vehicle_frame_no
    )
    SELECT
      rt.vehicle_frame_no,
      rt.source_policy_no,
      CAST(rt.expiry_date AS DATE) AS expiry_date,
      rt.coverage_combination,
      rt.is_renewed,
      rt.is_quoted,
      COALESCE(q.quote_insurance_grade, p.policy_insurance_grade, '') AS insurance_grade,
      COALESCE(q.quote_pricing_factor, p.policy_pricing_factor) AS pricing_factor,
      COALESCE(flow.next_insurer, '') AS loss_company
    FROM rt
    LEFT JOIN q_latest q ON q.vehicle_frame_no = rt.vehicle_frame_no
    LEFT JOIN policy_one p
      ON p.policy_no = rt.source_policy_no
     AND p.vehicle_frame_no = rt.vehicle_frame_no
    LEFT JOIN flow ON flow.vehicle_frame_no = rt.vehicle_frame_no
    """
    con = duckdb.connect(":memory:")
    return con.execute(sql, [config.start, config.end, config.quote_window_start]).fetchdf().to_dict("records")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def clean_number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    if out != out:
        return None
    return out


def date_to_epoch_ms(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
            try:
                value = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue
        else:
            try:
                value = datetime.fromisoformat(text[:10]).date()
            except ValueError:
                return None
    if isinstance(value, datetime):
        value = value.date()
    if not isinstance(value, date):
        return None
    # 企业微信 DATE_TIME 在前端按本地时区渲染；用 UTC noon 表示业务日期，
    # 避免 UTC midnight 在中国时区/浏览器转换时显示成前一天。
    dt = datetime.combine(value, time(12, 0), tzinfo=timezone.utc)
    return str(int(dt.timestamp() * 1000))


def yes_no(value: Any) -> str:
    return "是" if bool(value) else "否"


def select_value(value: Any) -> list[dict[str, str]] | None:
    text = clean_text(value)
    if not text:
        return None
    return [{"text": text}]


def reset_table_spec() -> None:
    CURRENT_FULL_FIELD_IDS.clear()
    CURRENT_FULL_FIELD_IDS.update(FULL_FIELD_IDS)
    CURRENT_FIELD_TYPES.clear()
    CURRENT_FIELD_TYPES.update(DEFAULT_FIELD_TYPES)


def field_id(key: str) -> str:
    return CURRENT_FULL_FIELD_IDS[key]


def schema_for(keys: Iterable[str]) -> dict[str, str]:
    return {field_id(key): KEY_LABELS[key] for key in keys}


def configured_keys(base_keys: Iterable[str]) -> list[str]:
    return list(base_keys)


def infer_field_type(sample_value: Any) -> str:
    if isinstance(sample_value, list):
        if sample_value and all(isinstance(item, dict) and "user_id" in item for item in sample_value):
            return "user"
        return "select"
    if isinstance(sample_value, (int, float)) and not isinstance(sample_value, bool):
        return "number"
    return "text"


def apply_table_schema_file(path: Path) -> None:
    """Override field IDs/types from the webhook sample JSON supplied by the sheet owner."""
    raw = json.loads(path.read_text(encoding="utf-8"))
    schema = raw.get("schema")
    if not isinstance(schema, dict):
        raise RuntimeError(f"schema 文件缺少 schema 对象: {path}")

    sample_records = raw.get("add_records") or []
    sample_values = {}
    if sample_records and isinstance(sample_records[0], dict):
        sample_values = sample_records[0].get("values") or {}

    reset_table_spec()
    for fid, label in schema.items():
        key = LABEL_TO_KEY.get(str(label).strip())
        if not key:
            continue
        CURRENT_FULL_FIELD_IDS[key] = str(fid)
        if fid in sample_values:
            CURRENT_FIELD_TYPES[key] = infer_field_type(sample_values[fid])

    CURRENT_FIELD_TYPES["expiry_date"] = "date"
    CURRENT_FIELD_TYPES["owner_user"] = "user"


def render_cell_value(key: str, raw: Any) -> Any | None:
    field_type = CURRENT_FIELD_TYPES.get(key, "text")
    if field_type == "select":
        return select_value(raw)
    if field_type == "number":
        return clean_number(raw)
    if field_type == "date":
        return date_to_epoch_ms(raw)
    if field_type == "user":
        text = clean_text(raw)
        if not text:
            return None
        # USER 字段必须是真实 user_id；导出的姓名不能安全回写。
        return [{"user_id": text}]
    text = clean_text(raw)
    return text or None


def build_update_values(row: dict[str, Any]) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for key, raw in (
        ("is_renewed", yes_no(row.get("is_renewed"))),
        ("is_quoted", yes_no(row.get("is_quoted"))),
        ("coverage_combination", row.get("coverage_combination")),
        ("insurance_grade", row.get("insurance_grade")),
        ("pricing_factor", row.get("pricing_factor")),
        ("loss_company", row.get("loss_company")),
    ):
        if key not in BASE_UPDATE_KEYS:
            continue
        if key not in CURRENT_FULL_FIELD_IDS:
            continue
        rendered = render_cell_value(key, raw)
        if rendered is not None:
            values[field_id(key)] = rendered
    return values


EXCEL_FIELD_TO_KEY = {
    "名单类型": "list_type",
    "投保险种": "insurance_combo",
    "车牌号": "plate_no",
    "车架号": "vehicle_frame_no",
    "保单到期时间": "expiry_date",
    "归属人": "owner",
    "归属人(人员)": "owner_user",
    "坐席域账号": "seat_account",
    "归属团队": "team",
    "车型": "vehicle_type",
    "险别组合": "coverage_combination",
}


def read_excel_rows(path: Path) -> list[dict[str, Any]]:
    # This WeCom-exported workbook reports wrong dimensions in openpyxl
    # read_only mode, so use normal mode to preserve all rows.
    workbook = load_workbook(path, read_only=False, data_only=True)
    rows: list[dict[str, Any]] = []
    for ws in workbook.worksheets:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        header_index = {h: idx for idx, h in enumerate(headers) if h}
        if "车架号" not in header_index:
            continue
        for excel_row_num, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row: dict[str, Any] = {"_sheet": ws.title, "_excel_row": excel_row_num}
            for label, key in EXCEL_FIELD_TO_KEY.items():
                idx = header_index.get(label)
                row[key] = values[idx] if idx is not None and idx < len(values) else None
            vin = clean_text(row.get("vehicle_frame_no"))
            if not vin:
                continue
            row["vehicle_frame_no"] = vin
            rows.append(row)
    return rows


def build_seed_values(excel_row: dict[str, Any], enrichment_row: dict[str, Any] | None) -> dict[str, Any]:
    values: dict[str, Any] = {}

    def put(key: str, raw: Any) -> None:
        if key == "owner_user":
            return
        rendered = render_cell_value(key, raw)
        if rendered is not None:
            values[field_id(key)] = rendered

    put("list_type", excel_row.get("list_type"))
    put("insurance_combo", excel_row.get("insurance_combo"))
    put("plate_no", excel_row.get("plate_no"))
    put("vehicle_frame_no", excel_row.get("vehicle_frame_no"))
    put("expiry_date", excel_row.get("expiry_date"))
    put("owner", excel_row.get("owner"))
    # 归属人(人员) 是 USER 字段；导出 Excel 只有姓名、没有 user_id。
    # 传空 user_id 容易触发接口错误，因此 seed 阶段不写该字段。
    put("seat_account", excel_row.get("seat_account"))
    put("team", excel_row.get("team"))
    put("vehicle_type", excel_row.get("vehicle_type"))
    put("coverage_combination", excel_row.get("coverage_combination"))

    if enrichment_row:
        values.update(build_update_values(enrichment_row))
    return values


def build_seed_records(excel_rows: list[dict[str, Any]], source_rows: list[dict[str, Any]]) -> dict[str, Any]:
    source_by_vin = {str(r.get("vehicle_frame_no")): r for r in source_rows if r.get("vehicle_frame_no")}
    records: list[dict[str, Any]] = []
    duplicate_vins: dict[str, int] = {}
    seen: set[str] = set()
    for row in excel_rows:
        vin = str(row.get("vehicle_frame_no"))
        if vin in seen:
            duplicate_vins[vin] = duplicate_vins.get(vin, 1) + 1
        seen.add(vin)
        values = build_seed_values(row, source_by_vin.get(vin))
        records.append({
            "values": values,
            "_vin": vin,
            "_sheet": row.get("_sheet"),
            "_excel_row": row.get("_excel_row"),
            "_matched_source": vin in source_by_vin,
        })
    return {
        "records": records,
        "duplicate_vins": duplicate_vins,
        "missing_source_vins": sorted({str(r.get("vehicle_frame_no")) for r in excel_rows} - set(source_by_vin)),
        "matched_source_count": sum(1 for r in records if r["_matched_source"]),
    }


def parse_vins(items: list[str] | None, vin_file: str | None = None) -> set[str] | None:
    vins: set[str] = set()
    for item in items or []:
        for part in str(item).replace("，", ",").split(","):
            vin = part.strip()
            if vin:
                vins.add(vin)
    if vin_file:
        text = Path(vin_file).expanduser().read_text(encoding="utf-8")
        for line in text.splitlines():
            for part in line.replace("，", ",").split(","):
                vin = part.strip()
                if vin:
                    vins.add(vin)
    return vins or None


def filter_rows_by_vin(rows: list[dict[str, Any]], vin_filter: set[str] | None) -> list[dict[str, Any]]:
    if not vin_filter:
        return rows
    return [row for row in rows if str(row.get("vehicle_frame_no") or "").strip() in vin_filter]


def build_inspection(rows: list[dict[str, Any]], requested_vins: set[str] | None) -> dict[str, Any]:
    found_vins = {str(row.get("vehicle_frame_no") or "").strip() for row in rows}
    inspection_rows = []
    for row in rows:
        inspection_rows.append({
            "vehicle_frame_no": row.get("vehicle_frame_no"),
            "source_policy_no": row.get("source_policy_no"),
            "expiry_date": row.get("expiry_date"),
            "fields": {
                "是否成交": yes_no(row.get("is_renewed")),
                "是否报价": yes_no(row.get("is_quoted")),
                "风险等级": clean_text(row.get("insurance_grade")),
                "自主系数": clean_number(row.get("pricing_factor")),
                "流失公司": clean_text(row.get("loss_company")),
            },
            "webhook_values": build_update_values(row),
        })
    return {
        "operation": "inspect_may_renewal_fields",
        "requested_vin_count": len(requested_vins or []),
        "matched_count": len(rows),
        "missing_requested_vins": sorted((requested_vins or set()) - found_vins),
        "schema": schema_for(configured_keys(BASE_UPDATE_KEYS)),
        "rows": inspection_rows,
    }


def payload_hash(values: dict[str, Any]) -> str:
    raw = json.dumps(values, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def build_plan(rows: list[dict[str, Any]], state: dict[str, Any], *, force: bool = False) -> dict[str, Any]:
    state_records = state.get("records", {})
    source_by_vin = {str(r.get("vehicle_frame_no")): r for r in rows if r.get("vehicle_frame_no")}
    update_records: list[dict[str, Any]] = []
    skipped_unchanged = 0

    for vin, row in source_by_vin.items():
        existing = state_records.get(vin)
        if not existing or not existing.get("record_id"):
            continue
        values = build_update_values(row)
        h = payload_hash(values)
        if not force and existing.get("payload_hash") == h:
            skipped_unchanged += 1
            continue
        update_records.append({
            "record_id": existing["record_id"],
            "values": values,
            "_vin": vin,
            "_payload_hash": h,
            "_source_policy_no": row.get("source_policy_no"),
        })

    return {
        "update_records": update_records,
        "source_by_vin": source_by_vin,
        "missing_in_state": sorted(set(source_by_vin) - set(state_records)),
        "missing_in_source": sorted(set(state_records) - set(source_by_vin)),
        "skipped_unchanged": skipped_unchanged,
    }


def plan_field_stats(update_records: list[dict[str, Any]]) -> dict[str, Any]:
    """Count how many planned updates touch each business field."""
    field_to_key = {field_id(key): key for key in BASE_UPDATE_KEYS if key in CURRENT_FULL_FIELD_IDS}
    counts = {key: 0 for key in BASE_UPDATE_KEYS}
    for record in update_records:
        for target_id in record.get("values", {}):
            key = field_to_key.get(target_id)
            if key:
                counts[key] = counts.get(key, 0) + 1
    return {
        "total_update_records": len(update_records),
        "field_counts": counts,
    }


def chunked(items: list[Any], size: int) -> Iterable[list[Any]]:
    for i in range(0, len(items), size):
        yield items[i:i + size]


def post_webhook(url: str, payload: dict[str, Any], timeout: int = 60) -> dict[str, Any]:
    # Reuse the established retry/error behavior from the main WeCom renewal engine.
    sys.path.insert(0, str(HERE))
    from sync_renewal_v2 import post_webhook as shared_post_webhook  # noqa: E402

    return shared_post_webhook(url, payload, timeout=timeout)


def write_log(instance_name: str, summary: dict[str, Any]) -> Path:
    log_dir = HERE / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = log_dir / f"{instance_name}_{ts}.json"
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path


def run_sync(config: SyncConfig, *, execute: bool) -> dict[str, Any]:
    state = load_state(config.state_path)
    if not state.get("records"):
        if not execute:
            rows = filter_rows_by_vin(fetch_source_rows(config), config.vin_filter)
            summary = {
                "operation": "sync_may_renewal_fields",
                "dry_run": True,
                "date_range": {"start": config.start, "end": config.end},
                "source_rows": len(rows),
                "state_records": 0,
                "to_update": 0,
                "state_required": True,
                "state_path": str(config.state_path),
                "message": "state 为空；真实更新前必须先执行 prime-state 建立 VIN→record_id 映射",
                "schema_field_ids": list(schema_for(configured_keys(BASE_UPDATE_KEYS)).keys()),
                "sample_source_rows": rows[:3],
            }
            log_path = write_log("renewal_may_5fields_dryrun", summary)
            summary["log_path"] = str(log_path)
            return summary
        raise RuntimeError(
            f"state 为空，不能更新现有表。请先执行 prime-state 建立 VIN→record_id 映射：{config.state_path}"
        )

    rows = filter_rows_by_vin(fetch_source_rows(config), config.vin_filter)
    plan = build_plan(rows, state, force=config.force)
    schema = schema_for(configured_keys(BASE_UPDATE_KEYS))
    summary: dict[str, Any] = {
        "operation": "sync_may_renewal_fields",
        "dry_run": not execute,
        "date_range": {"start": config.start, "end": config.end},
        "source_rows": len(rows),
        "state_records": len(state.get("records", {})),
        "to_update": len(plan["update_records"]),
        "skipped_unchanged": plan["skipped_unchanged"],
        "missing_in_state_count": len(plan["missing_in_state"]),
        "missing_in_state_sample": plan["missing_in_state"][:20],
        "missing_in_source_count": len(plan["missing_in_source"]),
        "missing_in_source_sample": plan["missing_in_source"][:20],
        "field_update_stats": plan_field_stats(plan["update_records"]),
        "schema_field_ids": list(schema.keys()),
        "sample_updates": [
            {"record_id": r["record_id"], "values": r["values"], "_vin": r["_vin"]}
            for r in plan["update_records"][:3]
        ],
    }
    if not execute:
        log_path = write_log("renewal_may_5fields_dryrun", summary)
        summary["log_path"] = str(log_path)
        return summary

    load_dotenv_local()
    webhook_url = config.webhook_url or os.environ.get(config.webhook_env)
    if not webhook_url:
        raise RuntimeError(f"缺少 webhook：请设置环境变量 {config.webhook_env} 或传 --webhook-url")

    batches: list[dict[str, Any]] = []
    for chunk in chunked(plan["update_records"], config.batch_size):
        payload_records = [{"record_id": r["record_id"], "values": r["values"]} for r in chunk]
        resp = post_webhook(webhook_url, {"schema": schema, "update_records": payload_records})
        if resp.get("errcode") != 0:
            raise RuntimeError(f"企业微信更新失败: {resp}")
        for item in chunk:
            state["records"][item["_vin"]] = {
                **state["records"].get(item["_vin"], {}),
                "record_id": item["record_id"],
                "payload_hash": item["_payload_hash"],
                "source_policy_no": item.get("_source_policy_no"),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
        save_state(config.state_path, state)
        batches.append({"op": "update", "sent": len(chunk), "errcode": resp.get("errcode")})

    summary["batches"] = batches
    summary["completed_at"] = datetime.now(timezone.utc).isoformat()
    log_path = write_log("renewal_may_5fields_sync", summary)
    summary["log_path"] = str(log_path)
    state["summary"] = {k: v for k, v in summary.items() if k != "sample_updates"}
    save_state(config.state_path, state)
    return summary


def run_seed_from_excel(config: SyncConfig, *, input_path: Path, execute: bool, allow_existing_state: bool = False) -> dict[str, Any]:
    state = load_state(config.state_path)
    if state.get("records") and not allow_existing_state:
        raise RuntimeError(
            f"state 已有 {len(state.get('records', {}))} 条记录，拒绝重复 seed。"
            "如确认目标表为空且要重建，请先备份/清理 state，或传 --allow-existing-state。"
        )

    excel_rows = read_excel_rows(input_path)
    vin_filter = {str(row.get("vehicle_frame_no")) for row in excel_rows if row.get("vehicle_frame_no")}
    source_rows = filter_rows_by_vin(fetch_source_rows(config), vin_filter)
    seed_plan = build_seed_records(excel_rows, source_rows)
    records = seed_plan["records"]
    schema = schema_for(configured_keys(BASE_SEED_KEYS))
    summary: dict[str, Any] = {
        "operation": "seed_from_excel",
        "dry_run": not execute,
        "input": str(input_path),
        "date_range": {"start": config.start, "end": config.end},
        "excel_rows": len(excel_rows),
        "unique_vins": len(vin_filter),
        "matched_source_count": seed_plan["matched_source_count"],
        "missing_source_count": len(seed_plan["missing_source_vins"]),
        "missing_source_sample": seed_plan["missing_source_vins"][:20],
        "duplicate_vin_count": len(seed_plan["duplicate_vins"]),
        "duplicate_vin_sample": sorted(seed_plan["duplicate_vins"])[:20],
        "to_add": len(records),
        "schema_field_ids": list(schema.keys()),
        "sample_records": [
            {k: v for k, v in record.items() if k != "values"} | {"values": record["values"]}
            for record in records[:3]
        ],
    }
    if not execute:
        log_path = write_log("renewal_may_jul_seed_dryrun", summary)
        summary["log_path"] = str(log_path)
        return summary

    load_dotenv_local()
    webhook_url = config.webhook_url or os.environ.get(config.webhook_env)
    if not webhook_url:
        raise RuntimeError(f"缺少 webhook：请设置环境变量 {config.webhook_env} 或传 --webhook-url")

    state_records = state.setdefault("records", {})
    batches: list[dict[str, Any]] = []
    for chunk in chunked(records, config.batch_size):
        payload_records = [{"values": record["values"]} for record in chunk]
        resp = post_webhook(webhook_url, {"schema": schema, "add_records": payload_records})
        if resp.get("errcode") != 0:
            raise RuntimeError(f"企业微信新增失败: {resp}")
        added = resp.get("add_records") or resp.get("data") or []
        if len(added) != len(chunk):
            raise RuntimeError(f"新增返回数量不一致: sent={len(chunk)} returned={len(added)}")
        for record, ret in zip(chunk, added):
            vin = record["_vin"]
            rid = ret.get("record_id") or ret.get("id")
            if not rid:
                raise RuntimeError(f"新增返回缺少 record_id: {ret}")
            state_records[vin] = {
                "record_id": rid,
                "payload_hash": payload_hash(record["values"]),
                "sheet": record.get("_sheet"),
                "excel_row": record.get("_excel_row"),
                "seeded_at": datetime.now(timezone.utc).isoformat(),
            }
        save_state(config.state_path, state)
        batches.append({"op": "add", "sent": len(chunk), "errcode": resp.get("errcode")})
        time_mod.sleep(60.0 / 3000 * len(chunk))

    summary["batches"] = batches
    summary["state_records_after"] = len(state_records)
    summary["completed_at"] = datetime.now(timezone.utc).isoformat()
    log_path = write_log("renewal_may_jul_seed", summary)
    summary["log_path"] = str(log_path)
    state["summary"] = {k: v for k, v in summary.items() if k != "sample_records"}
    save_state(config.state_path, state)
    return summary


def main() -> int:
    args = parse_args()
    try:
        if args.command == "prime-state":
            state_path = Path(args.state).expanduser().resolve()
            if args.records_json:
                response = load_records_json(Path(args.records_json).expanduser())
            else:
                if not args.sheet_id:
                    raise RuntimeError("--docid/--url 模式必须传 --sheet-id")
                response = get_records_with_wecom_cli(
                    docid=args.docid,
                    url=args.url,
                    sheet_id=args.sheet_id,
                    vin_field_id=args.vin_field_id,
                )
            summary = prime_state_from_records(
                response,
                state_path=state_path,
                vin_field_id=args.vin_field_id,
                vin_field_title=args.vin_field_title,
            )
        elif args.command == "inspect":
            vin_filter = parse_vins(args.vin, args.vin_file)
            if not vin_filter:
                raise RuntimeError("inspect 必须传 --vin 或 --vin-file，避免全量输出")
            config = SyncConfig(
                start=args.start,
                end=args.end,
                quote_window_start=args.quote_window_start,
                renewal_tracker_path=args.renewal_tracker_path,
                quotes_path=args.quotes_path,
                customer_flow_path=args.customer_flow_path,
                policy_glob=args.policy_glob,
                state_path=DEFAULT_STATE_PATH,
                webhook_env=DEFAULT_WEBHOOK_ENV,
                webhook_url=None,
                batch_size=100,
                force=False,
                vin_filter=vin_filter,
            )
            rows = filter_rows_by_vin(fetch_source_rows(config), vin_filter)
            summary = build_inspection(rows, vin_filter)
            if args.output:
                output = Path(args.output).expanduser()
                output.parent.mkdir(parents=True, exist_ok=True)
                output.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
                summary["output"] = str(output)
        elif args.command == "seed-from-excel":
            state_path = Path(args.state).expanduser().resolve()
            if args.table_schema_file:
                apply_table_schema_file(Path(args.table_schema_file).expanduser())
            BASE_UPDATE_KEYS[:] = SEED_UPDATE_KEYS
            config = SyncConfig(
                start=args.start,
                end=args.end,
                quote_window_start=args.quote_window_start,
                renewal_tracker_path=args.renewal_tracker_path,
                quotes_path=args.quotes_path,
                customer_flow_path=args.customer_flow_path,
                policy_glob=args.policy_glob,
                state_path=state_path,
                webhook_env=args.webhook_env,
                webhook_url=args.webhook_url,
                batch_size=args.batch_size,
                force=False,
                vin_filter=None,
            )
            summary = run_seed_from_excel(
                config,
                input_path=Path(args.input).expanduser(),
                execute=args.execute,
                allow_existing_state=args.allow_existing_state,
            )
            if args.output:
                output = Path(args.output).expanduser()
                output.parent.mkdir(parents=True, exist_ok=True)
                output.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
                summary["output"] = str(output)
        else:
            state_path = Path(args.state).expanduser().resolve()
            if getattr(args, "table_schema_file", None):
                apply_table_schema_file(Path(args.table_schema_file).expanduser())
            BASE_UPDATE_KEYS[:] = FIELD_SETS[getattr(args, "fields", "quote")]
            vin_filter = parse_vins(getattr(args, "vin", []), getattr(args, "vin_file", None))
            config = SyncConfig(
                start=args.start,
                end=args.end,
                quote_window_start=args.quote_window_start,
                renewal_tracker_path=args.renewal_tracker_path,
                quotes_path=args.quotes_path,
                customer_flow_path=args.customer_flow_path,
                policy_glob=args.policy_glob,
                state_path=state_path,
                webhook_env=args.webhook_env,
                webhook_url=args.webhook_url,
                batch_size=args.batch_size,
                force=args.force,
                vin_filter=vin_filter,
            )
            summary = run_sync(config, execute=args.execute)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
