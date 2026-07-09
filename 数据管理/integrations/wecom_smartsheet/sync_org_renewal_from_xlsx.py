#!/usr/bin/env python3
"""Batch-sync organization renewal tracker sheets from a local xlsx registry.

The registry keeps webhook URLs outside git. This wrapper reuses the existing
per-organization YAML instances and VIN->record_id state files, so reruns are
incremental by default.
"""

from __future__ import annotations

# --- bootstrap: load .env.local for standalone python3 invocation (see _env.py) ---
import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).resolve().parent))
import _env as _env  # noqa: F401,E402  module-level load_dotenv_local() runs on import
del _sys, _Path
# --- end bootstrap ---

import argparse
import json
import os
import sys
from dataclasses import replace
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parents[2]
DEFAULT_XLSX = (
    Path.home()
    / "Library"
    / "Mobile Documents"
    / "com~apple~CloudDocs"
    / "续保追踪表链接与意见反馈.xlsx"
)

sys.path.insert(0, str(HERE))
from sync_renewal_v2 import (  # noqa: E402
    InstanceConfig,
    build_record,
    build_source_rows,
    load_field_registry,
    load_instance,
    load_state,
    payload_hash,
    resolve_state_path,
    run_sync,
    save_state,
)


BRANCH_CODE = "SC"  # 本脚本同步「四川」续保追踪表；slug 映射见 org-slugs.json 同名 branch_code 键
ORG_SLUG_CONFIG = HERE / "org-slugs.json"


def _load_org_slugs(branch_code: str) -> dict[str, str]:
    """从省份化 SSOT（org-slugs.json）读取指定省份的「机构中文名→拼音 slug」映射。

    多省硬编码债治理（BACKLOG 2026-07-07-claude-cfaf91）：slug 曾以两份平行 dict
    硬编码在本文件，山西 / 新省份上线时零复用。抽到数据后，新省份只需在 JSON 加一个
    branch_code 顶层键，脚本零改动。fail-closed：配置缺失 / 省份无映射 / 空表一律报错
    中止，禁止静默回落空映射（空映射会让每个机构都命中「未配置机构」误报，掩盖真因）。
    """
    if not ORG_SLUG_CONFIG.exists():
        raise RuntimeError(f"缺少机构 slug 配置文件：{ORG_SLUG_CONFIG}")
    data = json.loads(ORG_SLUG_CONFIG.read_text(encoding="utf-8"))
    slugs = data.get(branch_code)
    if not slugs:
        raise RuntimeError(
            f"org-slugs.json 缺少省份 {branch_code!r} 的机构 slug 映射（fail-closed，禁止静默空表）"
        )
    return dict(slugs)


# 机构中文名 → 拼音 slug（续保实例 YAML 文件名用，如 sichuan_gaoxin_2025_may_jul.yaml）
ORG_SLUGS = _load_org_slugs(BRANCH_CODE)
# 机构中文名 → webhook 环境变量后缀（WECOM_SMARTSHEET_WEBHOOK_<后缀>）。恒等于 slug 的
# 大写形态，故由 slug 单一来源派生，消除两份 dict 手工同步的漂移风险（不变量由单测锁死）。
ORG_ENVS = {org: slug.upper() for org, slug in ORG_SLUGS.items()}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="从 xlsx webhook 登记表批量同步机构续保追踪智能表格"
    )
    parser.add_argument(
        "--xlsx",
        default=str(DEFAULT_XLSX),
        help=f"机构/webhook 登记表路径，默认：{DEFAULT_XLSX}",
    )
    parser.add_argument(
        "--execute",
        action="store_true",
        help="真实调用 webhook；默认只 dry-run",
    )
    parser.add_argument(
        "--prime-hashes",
        action="store_true",
        help="只根据当前本地数据补齐 state.payload_hash，不调用 webhook；仅在确认表格已同步后使用",
    )
    parser.add_argument(
        "--org",
        help="只同步指定机构，多个用英文逗号分隔，例如：天府,高新,新都",
    )
    parser.add_argument("--batch-size", type=int, default=100)
    parser.add_argument(
        "--i-checked-wecom-rows",
        action="store_true",
        dest="i_checked_wecom_rows",
        help=(
            "RED LINE 闸覆盖开关：state 空 / to_add 占比过高 / 全 add 无 update 等"
            "危险信号触发时，--execute 默认拒绝执行（防止重复 add）。仅在你已**亲自**"
            "去企微表点查当前行数，并确认与 preflight banner 的 state.records_count 或 0"
            "吻合时，才能加此开关放行。详见 [[project_wecom_org_renewal_first_real_run_dup]]。"
        ),
    )
    return parser.parse_args()


def read_registry(path: Path, org_filter: set[str] | None) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        org_idx = header.index("机构")
        webhook_idx = header.index("webhook 地址")
        link_idx = header.index("智能表链接")
    except ValueError as exc:
        raise RuntimeError("xlsx 必须包含列：机构、webhook 地址、智能表链接") from exc

    rows: list[dict[str, str]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        org = str(values[org_idx] or "").strip()
        webhook = str(values[webhook_idx] or "").strip()
        link = str(values[link_idx] or "").strip()
        if not org or org == "模板" or not webhook:
            continue
        if org_filter and org not in org_filter:
            continue
        if org not in ORG_SLUGS:
            raise RuntimeError(f"登记表中存在未配置机构：{org}")
        if not webhook.startswith(
            "https://qyapi.weixin.qq.com/cgi-bin/wedoc/smartsheet/webhook?key="
        ):
            raise RuntimeError(f"{org} 的 webhook 地址格式异常")
        rows.append({"org": org, "webhook": webhook, "link": link})
    return rows


def resolve_instance_path(org: str) -> Path:
    slug = ORG_SLUGS[org]
    enabled = HERE / "instances" / f"sichuan_{slug}_2025_may_jul.yaml"
    disabled = HERE / "instances" / f"sichuan_{slug}_2025_may_jul.yaml.disabled"
    if enabled.exists():
        return enabled
    if disabled.exists():
        return disabled
    raise RuntimeError(f"缺少机构实例 YAML：{org}")


def load_fields(instance: InstanceConfig):
    if instance.field_registry_path:
        registry_path = Path(instance.field_registry_path)
        if not registry_path.is_absolute():
            registry_path = HERE / registry_path
    else:
        registry_path = HERE / "field_registry.yaml"
    registry = load_field_registry(registry_path)
    missing = [name for name in instance.fields_enabled if name not in registry]
    if missing:
        raise RuntimeError(f"{instance.instance_name} 启用了未注册字段：{missing}")
    return [registry[name] for name in instance.fields_enabled]


def compact_summary(org: str, link: str, summary: dict[str, Any]) -> dict[str, Any]:
    return {
        "org": org,
        "link": link,
        "instance": summary.get("instance_name"),
        "dry_run": summary.get("dry_run"),
        "source_rows": summary.get("source_rows"),
        "to_add": summary.get("to_add"),
        "to_update": summary.get("to_update"),
        "changed_rows": summary.get("changed_rows"),
        "premium_sum": summary.get("premium_sum"),
        "renewal_rate": summary.get("renewal_rate"),
        "add_premium_sum": summary.get("add_premium_sum"),
        "add_renewed_count": summary.get("add_renewed_count"),
        "add_renewal_rate": summary.get("add_renewal_rate"),
        "update_premium_sum": summary.get("update_premium_sum"),
        "update_renewed_count": summary.get("update_renewed_count"),
        "update_renewal_rate": summary.get("update_renewal_rate"),
        "changed_premium_sum": summary.get("changed_premium_sum"),
        "changed_renewed_count": summary.get("changed_renewed_count"),
        "changed_renewal_rate": summary.get("changed_renewal_rate"),
        "renewed_count": summary.get("renewed_count"),
        "quoted_count": summary.get("quoted_count"),
        "missing_vins_count": summary.get("missing_vins_count"),
        "unmatched_salesmen_count": summary.get("unmatched_salesmen_count"),
        "update_failure_count": summary.get("update_failure_count"),
        "state_records_after": summary.get("state_records_after"),
        "hash_primed_rows": summary.get("hash_primed_rows"),
        "log_path": summary.get("log_path"),
    }


def prime_state_hashes(instance: InstanceConfig, fields) -> dict[str, Any]:
    rows, audit = build_source_rows(instance)
    state_path = resolve_state_path(instance)
    state = load_state(state_path)
    records = state.setdefault("records", {})
    matched = 0
    missing_record_id = 0
    for row in rows:
        vin = str(row.get("vehicle_frame_no"))
        existing = records.get(vin)
        if not existing or not existing.get("record_id"):
            missing_record_id += 1
            continue
        existing["payload_hash"] = payload_hash(build_record(row, fields, unmatched_set=set()))
        matched += 1
    state["summary"] = {
        **state.get("summary", {}),
        "hash_primed": True,
        "hash_primed_rows": matched,
        "hash_missing_record_id": missing_record_id,
        "duplicate_commercial_vin_count": audit["duplicate_commercial_vin_count"],
    }
    save_state(state_path, state)
    return {
        "dry_run": True,
        "instance_name": instance.instance_name,
        "source_rows": len(rows),
        "to_add": missing_record_id,
        "to_update": 0,
        "missing_vins_count": len(set(records) - {str(r.get("vehicle_frame_no")) for r in rows}),
        "unmatched_salesmen_count": None,
        "state_records_after": len(records),
        "hash_primed_rows": matched,
        "log_path": None,
    }


def main() -> int:
    args = parse_args()
    org_filter = {item.strip() for item in args.org.split(",")} if args.org else None
    registry_rows = read_registry(Path(args.xlsx).expanduser(), org_filter)
    if not registry_rows:
        raise RuntimeError("没有找到可同步的机构 webhook 行")

    results: list[dict[str, Any]] = []
    failed = False
    for row in registry_rows:
        org = row["org"]
        env_name = f"WECOM_SMARTSHEET_WEBHOOK_{ORG_ENVS[org]}"
        old_value = os.environ.get(env_name)
        os.environ[env_name] = row["webhook"]
        try:
            instance = load_instance(resolve_instance_path(org))
            if args.batch_size:
                instance = replace(instance, batch_size=args.batch_size)
            fields = load_fields(instance)

            if args.prime_hashes:
                summary = prime_state_hashes(instance, fields)
            else:
                # RED LINE preflight：永远先跑 dry-run 拿 plan，打 banner + 跑 gate；
                # 通过且 args.execute 才真写入。dry-run 不调 webhook，只 xlsx 解析。
                from _safety import (  # 局部 import 避免循环
                    evaluate_gate,
                    print_preflight_banner,
                    must_check_wecom_rows_hint,
                )

                state_path = resolve_state_path(instance)
                state_check = load_state(state_path) if state_path.exists() else {"records": {}}
                state_count = len(state_check.get("records") or {})

                plan = run_sync(instance, fields, dry_run=True)
                gate = print_preflight_banner(
                    label=f"{org} 续保追踪表",
                    state_count=state_count,
                    source_rows=plan.get("source_rows") or 0,
                    to_add=plan.get("to_add") or 0,
                    to_update=plan.get("to_update") or 0,
                )

                if not gate.ok and args.execute and not args.i_checked_wecom_rows:
                    raise RuntimeError(
                        gate.message + must_check_wecom_rows_hint(org, state_count)
                    )

                if args.execute:
                    summary = run_sync(instance, fields, dry_run=False)
                else:
                    summary = plan
            results.append(compact_summary(org, row["link"], summary))
        except Exception as exc:  # keep later orgs visible in one run
            failed = True
            results.append({"org": org, "link": row["link"], "error": str(exc)[:500]})
        finally:
            if old_value is None:
                os.environ.pop(env_name, None)
            else:
                os.environ[env_name] = old_value

    total = {
        "execute": args.execute,
        "org_count": len(results),
        "source_rows": sum(item.get("source_rows") or 0 for item in results),
        "to_add": sum(item.get("to_add") or 0 for item in results),
        "to_update": sum(item.get("to_update") or 0 for item in results),
        "changed_rows": sum(item.get("changed_rows") or 0 for item in results),
        "changed_premium_sum": round(sum(item.get("changed_premium_sum") or 0 for item in results), 2),
        "changed_renewed_count": sum(item.get("changed_renewed_count") or 0 for item in results),
        "missing_vins_count": sum(item.get("missing_vins_count") or 0 for item in results),
        "failed_count": sum(1 for item in results if item.get("error")),
    }
    total["changed_renewal_rate"] = (
        round(total["changed_renewed_count"] / total["changed_rows"], 4)
        if total["changed_rows"]
        else None
    )
    print(json.dumps({"total": total, "results": results}, ensure_ascii=False, indent=2))
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
